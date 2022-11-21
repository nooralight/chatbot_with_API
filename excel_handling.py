import openpyxl

def excel_return(path):
    #path = "testing.xlsx"
    # To open the workbook 
    # workbook object is created 
    wb_obj = openpyxl.load_workbook(path)
    # Get workbook active sheet object 
    # from the active attribute 
    sheet_obj = wb_obj.active
    # Cell objects also have a row, column, 
    # and coordinate attributes that provide 
    # location information for the cell. 
    # Note: The first row or 
    # column integer is 1, not 0. 
    # Cell object is created by using 
    # sheet object's cell() method. 
    row = sheet_obj.max_row
    column = sheet_obj.max_column
    #cell_obj = sheet_obj.cell(row = 2, column = 1)
    # Print value of cell object 
    # using the value attribute
    info_dict = []
    
    for i in range(row):
        
        
        info_dict.append({"name":sheet_obj.cell(row = i+1, column = 1).value,"category":sheet_obj.cell(row = i+1, column = 2).value})
    return info_dict

print(excel_return("testing.xlsx"))